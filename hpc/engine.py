"""Analysis engine for the High Performance Diagnostic Tool. PACE framework."""
from __future__ import annotations
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any
import numpy as np
import pandas as pd
from .config_loader import HPCConfig, PILLARS


@dataclass
class PillarResult:
    pillar: str; mean: float; status: str
    gap_vs_company: float; interpretation: str; std: float = 0.0


@dataclass
class DistributionStat:
    """Per-element distribution + variance (robust, plain-float)."""
    pillar: str
    mean: float
    std: float
    variance: float
    pct_low: float
    pct_mid: float
    pct_high: float
    polarised: bool
    variance_label: str
    statement: str


@dataclass
class DepartmentResult:
    department: str; n_respondents: int
    pillar_means: dict[str, float]
    overall: float; imbalance: float; balance_label: str
    classification: str; downgraded: bool; raw_classification: str
    pillar_results: list[PillarResult] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class AnalysisResult:
    focus: DepartmentResult
    company_pillar_means: dict[str, float]
    company_overall: float; company_n: int
    all_departments: pd.DataFrame
    imbalance_by_dept: pd.Series
    correlation: pd.DataFrame
    insights: list[dict[str, str]]
    recommendations: list[dict[str, Any]]
    distribution: list[DistributionStat] = field(default_factory=list)
    tailored_recommendations: list[dict[str, Any]] = field(default_factory=list)
    focus_dept_name: str = ""

    # ---- Backwards-compat: older code referenced analysis.polarisation ----
    @property
    def polarisation(self) -> list[DistributionStat]:
        return self.distribution


def classify_score(score: float, cfg: HPCConfig) -> str:
    if score >= cfg.band_performing_max + 0.001: return "High Performance"
    if score >= cfg.band_balanced_max + 0.001: return "Performing"
    if score >= cfg.band_dysfunctional_max + 0.001: return "Developing"
    return "At Risk"


def balance_label(gap: float, cfg: HPCConfig) -> str:
    if gap <= cfg.imbalance_wellbalanced_max: return "Well balanced"
    if gap <= cfg.imbalance_moderate_max: return "Moderate imbalance"
    return "Significant imbalance"


def apply_imbalance_downgrade(base: str, gap: float, cfg: HPCConfig) -> tuple[str, bool]:
    if gap > cfg.imbalance_moderate_max:
        order = ["At Risk", "Developing", "Performing", "High Performance"]
        idx = order.index(base) if base in order else len(order) - 1
        if idx > 0:
            return order[idx - 1], True
    return base, False


PILLAR_INTERPRETATIONS = {
    "Purpose":       {"strong": "Strategic clarity and shared direction are a genuine strength.",
                      "middle": "Direction is understood, but line-of-sight to outcomes can be sharpened.",
                      "weak":   "Weak strategic clarity — teams are unsure why they exist or what to prioritise."},
    "Alliance":      {"strong": "Leadership, trust and stakeholder relationships are a real strength.",
                      "middle": "Relationships and communication are adequate but not distinctive.",
                      "weak":   "Leadership visibility and cross-functional trust need active rebuilding."},
    "Collaboration": {"strong": "Ways of working, systems and decision-making enable performance.",
                      "middle": "Ways of working are functional but generate friction under pressure.",
                      "weak":   "Primary drag on performance — systems, tools and hand-offs create friction."},
    "Excellence":    {"strong": "Learning, feedback and continuous improvement are the differentiator.",
                      "middle": "Learning and improvement are present but not yet systematic.",
                      "weak":   "Weak learning muscle — improvement is ad hoc rather than a habit."},
}


def _interp(p: str, s: float) -> str:
    key = "weak" if s < 5.5 else ("strong" if s >= 7.0 else "middle")
    return PILLAR_INTERPRETATIONS[p][key]


_CLOSING = {
    "strength": {"Purpose": "Keep reinforcing the vision at every all-hands so it stays a strength.",
                 "Alliance": "Use this trust as the platform for tougher cross-team commitments.",
                 "Collaboration": "Document these workflows as the standard for other teams to reuse.",
                 "Excellence": "Turn this learning habit into a visible playbook others can copy."},
    "concern": {"Purpose": "Re-anchor the team on the top three priorities without delay.",
                "Alliance": "Prioritise rebuilding leadership visibility and stakeholder trust.",
                "Collaboration": "Target the biggest hand-off bottleneck first for a quick win.",
                "Excellence": "Introduce a lightweight, recurring learning and feedback rhythm."},
    "split": {"Purpose": "Find out which sub-teams feel disconnected from the strategy before acting.",
              "Alliance": "Identify which relationships are strained rather than applying a blanket fix.",
              "Collaboration": "Map which workflows work and which don't before standardising.",
              "Excellence": "Learn what the high-rating group does differently and spread it."},
    "scattered": {"Purpose": "Segment by role to locate where purpose is being lost.",
                  "Alliance": "Break the score down by team to surface pockets of low trust.",
                  "Collaboration": "Analyse by workflow to find the specific friction points.",
                  "Excellence": "Split by function to see where learning support is uneven."},
    "settled": {"Purpose": "A single average is reliable here — monitor at the next wave.",
                "Alliance": "Views are consistent — hold the current relationship rhythm.",
                "Collaboration": "Process experience is stable — watch for drift over time.",
                "Excellence": "Learning sentiment is steady — maintain the current investment."},
}


def _distribution_for_pillar(pillar: str, scores: np.ndarray) -> DistributionStat:
    scores = np.asarray(scores, dtype=float)
    scores = scores[~np.isnan(scores)]
    if len(scores) == 0:
        return DistributionStat(pillar, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, False, "No data", "No data available.")
    mean = float(np.mean(scores)); std = float(np.std(scores, ddof=1) if len(scores) > 1 else 0.0)
    var = float(std ** 2); n = len(scores)
    pct_low = float(100.0 * np.sum(scores <= 4) / n)
    pct_mid = float(100.0 * np.sum((scores >= 5) & (scores <= 6)) / n)
    pct_high = float(100.0 * np.sum(scores >= 7) / n)
    if std >= 2.0: vlabel = "High variance"
    elif std >= 1.3: vlabel = "Moderate variance"
    else: vlabel = "Low variance"
    split = bool(pct_low >= 25 and pct_high >= 25)
    c = _CLOSING
    if split:
        stmt = (f"{pillar} shows a <b>split view</b> — {pct_low:.0f}% rate it low (1-4) while {pct_high:.0f}% "
                f"rate it high (7-10). {vlabel} (SD={std:.2f}) means a single average hides real disagreement. "
                f"{c['split'][pillar]}")
    elif pct_high >= 60:
        stmt = (f"{pillar} is a <b>broad, agreed strength</b> — {pct_high:.0f}% rate it high (7-10) with "
                f"{vlabel.lower()} (SD={std:.2f}). {c['strength'][pillar]}")
    elif pct_low >= 45:
        stmt = (f"{pillar} is a <b>shared concern</b> — {pct_low:.0f}% rate it low (1-4), {vlabel.lower()} "
                f"(SD={std:.2f}); consistent, not an outlier. {c['concern'][pillar]}")
    elif std >= 2.0:
        stmt = (f"{pillar} shows <b>scattered views</b> (SD={std:.2f}) with no clear majority "
                f"({pct_low:.0f}% low, {pct_mid:.0f}% mid, {pct_high:.0f}% high). {c['scattered'][pillar]}")
    else:
        stmt = (f"{pillar} shows a <b>settled, consensual view</b> (SD={std:.2f}); {pct_high:.0f}% high, "
                f"{pct_low:.0f}% low. {c['settled'][pillar]}")
    return DistributionStat(pillar, round(mean, 2), round(std, 2), round(var, 2),
                            round(pct_low, 1), round(pct_mid, 1), round(pct_high, 1),
                            split, vlabel, stmt)


def analyze(responses: pd.DataFrame, focus_dept: str, cfg: HPCConfig) -> AnalysisResult:
    required = {"Department", "Pillar", "Score", "Submission ID"}
    if required - set(responses.columns):
        raise ValueError(f"Missing columns: {required - set(responses.columns)}")
    df = responses.copy()
    df["Score"] = pd.to_numeric(df["Score"], errors="coerce")
    df = df.dropna(subset=["Score"])
    if len(df) == 0:
        raise ValueError("No scored responses available to analyse.")

    dept_pillar = df.groupby(["Department", "Pillar"])["Score"].mean().unstack().reindex(columns=PILLARS)
    company_pillar = df.groupby("Pillar")["Score"].mean().reindex(PILLARS)
    company_overall = float(company_pillar.mean())
    company_n = int(df["Submission ID"].nunique())
    dept_overall = dept_pillar.mean(axis=1)
    imbalance_by_dept = (dept_pillar.max(axis=1) - dept_pillar.min(axis=1)).sort_values()
    all_depts = dept_pillar.copy()
    all_depts["Overall"] = dept_overall
    all_depts["Imbalance"] = dept_pillar.max(axis=1) - dept_pillar.min(axis=1)
    all_depts["N respondents"] = df.groupby("Department")["Submission ID"].nunique()
    sub_pillar = df.groupby(["Submission ID", "Pillar"])["Score"].mean().unstack().reindex(columns=PILLARS)
    correlation = sub_pillar.corr()

    if focus_dept == "__ALL__" or focus_dept not in dept_pillar.index:
        fps = company_pillar.copy(); focus_overall = company_overall
        n_focus = company_n; focus_name = "Company-wide"
        focus_std = df.groupby("Pillar")["Score"].std().reindex(PILLARS)
        pol_df = df
    else:
        fps = dept_pillar.loc[focus_dept]
        focus_overall = float(dept_overall.loc[focus_dept])
        n_focus = int(df[df.Department == focus_dept]["Submission ID"].nunique())
        focus_name = focus_dept
        focus_std = df[df.Department == focus_dept].groupby("Pillar")["Score"].std().reindex(PILLARS)
        pol_df = df[df.Department == focus_dept]

    focus_means = {p: float(fps[p]) for p in PILLARS}
    focus_imb = float(fps.max() - fps.min())
    raw_class = classify_score(focus_overall, cfg)
    final_class, downgraded = apply_imbalance_downgrade(raw_class, focus_imb, cfg)
    bal = balance_label(focus_imb, cfg)

    prs = [PillarResult(pillar=p, mean=focus_means[p], status=classify_score(focus_means[p], cfg),
                        gap_vs_company=focus_means[p] - float(company_pillar[p]),
                        interpretation=_interp(p, focus_means[p]),
                        std=(float(focus_std[p]) if not pd.isna(focus_std[p]) else 0.0)) for p in PILLARS]

    warnings = []
    if n_focus < cfg.min_responses_dept and focus_name != "Company-wide":
        warnings.append(f"Interpret with caution: {n_focus} < {cfg.min_responses_dept} responses.")

    focus_result = DepartmentResult(
        department=focus_name, n_respondents=n_focus, pillar_means=focus_means,
        overall=focus_overall, imbalance=focus_imb, balance_label=bal,
        classification=final_class, downgraded=downgraded, raw_classification=raw_class,
        pillar_results=prs, warnings=warnings,
    )

    distribution = []
    for p in PILLARS:
        s = pol_df[pol_df.Pillar == p]["Score"].to_numpy(dtype=float)
        distribution.append(_distribution_for_pillar(p, s))

    insights = _gen_insights(focus_result, dept_pillar, imbalance_by_dept, company_pillar, cfg)
    recs = _gen_recs(focus_result)
    tailored = _gen_tailored(focus_result, distribution, company_pillar)

    return AnalysisResult(
        focus=focus_result,
        company_pillar_means={p: float(company_pillar[p]) for p in PILLARS},
        company_overall=company_overall, company_n=company_n,
        all_departments=all_depts.round(3),
        imbalance_by_dept=imbalance_by_dept,
        correlation=correlation.round(3),
        insights=insights, recommendations=recs,
        distribution=distribution, tailored_recommendations=tailored,
        focus_dept_name=focus_name,
    )


def _gen_insights(focus, dept_pillar, imb_by_dept, company_pillar, cfg):
    ins = []
    means = focus.pillar_means
    strong = max(means, key=means.get); weak = min(means, key=means.get)
    gv = {p: means[p] - float(company_pillar[p]) for p in PILLARS}
    ins.append({"label": "Strongest element", "text": f"{strong} ({means[strong]:.2f}), {gv[strong]:+.2f} vs. company."})
    ins.append({"label": "Weakest element", "text": f"{weak} ({means[weak]:.2f}), {gv[weak]:+.2f} vs. company."})
    largest = min(gv, key=gv.get)
    ins.append({"label": "Largest gap vs. company", "text": f"{largest} ({gv[largest]:+.2f})."})
    if focus.imbalance > cfg.imbalance_moderate_max:
        ins.append({"label": "Imbalance", "text": f"Gap {focus.imbalance:.2f} exceeds threshold — classification downgraded."})
    elif focus.imbalance > cfg.imbalance_wellbalanced_max:
        ins.append({"label": "Imbalance", "text": f"Gap {focus.imbalance:.2f} — moderate; watch item."})
    else:
        ins.append({"label": "Balance", "text": f"Gap {focus.imbalance:.2f} — well balanced across elements."})
    if len(dept_pillar) > 1:
        ins.append({"label": "Most balanced dept", "text": f"{imb_by_dept.index[0]} (gap {imb_by_dept.iloc[0]:.2f})."})
        ins.append({"label": "Most imbalanced dept", "text": f"{imb_by_dept.index[-1]} (gap {imb_by_dept.iloc[-1]:.2f})."})
    return ins


ACTION_LIBRARY = {
    "Purpose": [
        ("Refresh the strategy cascade — clarify the top 3 priorities at every all-hands.", "High",
         "Rebuilds line-of-sight; lifts Purpose by 0.5-0.8.", "Head of Dept + Strategy Lead", "0-60 days"),
        ("Run customer-impact story sessions to reconnect work to outcomes.", "Medium",
         "Strengthens shared purpose.", "Head of Dept + CX Lead", "30-90 days"),
    ],
    "Alliance": [
        ("Introduce a structured joint-planning cadence with your top-5 stakeholders.", "High",
         "Rebuilds trust; lifts Alliance by 0.5+.", "Head of Dept + BU heads", "0-60 days"),
        ("Run 'Leader Listening' sessions and publish a 30-60-90 response plan.", "High",
         "Signals responsiveness; raises leadership visibility.", "Head of Dept + Leadership", "0-45 days"),
    ],
    "Collaboration": [
        ("Launch a 60-day workflow rescue on the top-3 cross-team bottlenecks.", "Critical",
         "Moves Collaboration by 0.6-1.0; cuts hand-off delays.", "Head of Dept + COO sponsor", "0-90 days"),
        ("Clarify decision rights and rationalise governance forums.", "High",
         "Improves decision speed and role clarity.", "Head of Dept + Chief of Staff", "30-90 days"),
    ],
    "Excellence": [
        ("Publish a capability roadmap and safeguard the L&D budget.", "High",
         "Preserves the differentiator under cost pressure.", "Head of Dept + P&C BP", "30-90 days"),
        ("Introduce a monthly feedback ritual (peer + upward + downward).", "Medium",
         "Builds a systematic feedback culture.", "Head of Dept + Leadership", "0-60 days"),
    ],
}


def _gen_recs(focus):
    ordered = sorted(focus.pillar_means.items(), key=lambda x: x[1])
    recs = []
    for i, (pillar, mean) in enumerate(ordered):
        actions = ACTION_LIBRARY.get(pillar, [])
        n = 2 if i < 2 else 1
        for action, priority, impact, owner, timeline in actions[:n]:
            if i == 0 and priority == "High":
                priority = "Critical" if pillar == "Collaboration" or mean < 4.0 else "High"
            recs.append({"Action": action, "Pillar": pillar, "Priority": priority,
                         "Expected Impact": impact, "Owner": owner, "Timeline": timeline})
    recs.append({"Action": "Re-run the PACE diagnostic in 6 months and compare movement.", "Pillar": "All",
                 "Priority": "Medium", "Expected Impact": "Creates accountability; validates ROI.",
                 "Owner": "OD / Cathay Academy", "Timeline": "6 months"})
    return recs


def _gen_tailored(focus, distribution, company_pillar):
    dist_by = {p.pillar: p for p in distribution}
    means = focus.pillar_means
    ordered = sorted(means.items(), key=lambda x: x[1])
    out = []
    for pillar, mean in ordered:
        d = dist_by.get(pillar)
        gap = mean - float(company_pillar[pillar])
        if d and d.polarised:
            action = (f"Run a targeted listening session on {pillar} to understand why the team is split "
                      f"({d.pct_low:.0f}% low vs {d.pct_high:.0f}% high) before rolling out a single fix.")
            rationale = f"Split view (SD={d.std:.2f}) — a blanket action risks alienating one half."
            priority = "High"
        elif mean < 4.0:
            action = f"Treat {pillar} as an At-Risk element — stand up a recovery plan with executive sponsorship now."
            rationale = f"Very low mean ({mean:.2f}); systemic issue."
            priority = "Critical"
        elif mean < 5.5:
            action = f"Prioritise a focused improvement sprint on {pillar} — the clearest drag on performance."
            rationale = f"Low mean ({mean:.2f}); {d.variance_label.lower() if d else ''}."
            priority = "High"
        elif mean >= 7.0 and d and d.pct_high >= 55:
            action = f"Codify what makes {pillar} work into a repeatable playbook for other teams."
            rationale = f"Consistent strength ({mean:.2f}, {d.pct_high:.0f}% high) worth scaling."
            priority = "Medium"
        elif d and d.std >= 2.0:
            action = f"Segment {pillar} by role/sub-team — scattered views hide pockets of risk."
            rationale = f"High variance (SD={d.std:.2f}) with no clear majority."
            priority = "Medium"
        else:
            action = f"Maintain steady improvement on {pillar}; monitor at the next wave."
            rationale = f"Stable mid-range view ({mean:.2f})."
            priority = "Low"
        out.append({"Element": pillar, "Mean": f"{mean:.2f}", "Gap": f"{gap:+.2f}",
                    "Recommendation": action, "Why": rationale, "Priority": priority})
    return out


def load_responses(path: str) -> pd.DataFrame:
    if str(path).lower().endswith(".csv"):
        df = pd.read_csv(path)
    else:
        xl = pd.ExcelFile(path)
        df = pd.read_excel(path, sheet_name="Responses" if "Responses" in xl.sheet_names else 0)
    if {"Submission ID", "Department", "Pillar", "Score"} - set(df.columns):
        raise ValueError("Response file missing required columns")
    return df


def append_submission(path: str, submission_id: str, department: str,
                      respondent_id: str, answers: dict) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(path)
    if "Responses" not in wb.sheetnames:
        raise ValueError("No 'Responses' sheet.")
    ws = wb["Responses"]
    row = ws.max_row + 1
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for qid, (qtext, pillar, score) in answers.items():
        ws.cell(row=row, column=1, value=submission_id)
        ws.cell(row=row, column=2, value=ts)
        ws.cell(row=row, column=3, value=department)
        ws.cell(row=row, column=4, value=respondent_id or "")
        ws.cell(row=row, column=5, value=qid)
        ws.cell(row=row, column=6, value=qtext)
        ws.cell(row=row, column=7, value=pillar)
        ws.cell(row=row, column=8, value=int(score))
        row += 1
    wb.save(path)
