"""PACE gamification engine — running-themed stages, badges, action plans."""
from __future__ import annotations
import json
import random
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

ACTIVATION_THRESHOLD = 0.70

PACE_STAGES = [
    "Not Started", "Warm-up Exercise", "Training", "Reflect", "Implement Change",
    "Reflect Again", "Training - Tempo", "Race up your PACE", "Finish Line",
]
STAGE_TO_INDEX = {s: i for i, s in enumerate(PACE_STAGES)}
STAGE_DISPLAY = {
    "Not Started": "Not Started", "Warm-up Exercise": "Warm-up Exercise", "Training": "Training",
    "Reflect": "Reflect", "Implement Change": "Implement Change", "Reflect Again": "Reflect",
    "Training - Tempo": "Training - Tempo", "Race up your PACE": "Race 'up your PACE'", "Finish Line": "Finish Line",
}
BADGES = [
    ("Warmed Up", "Survey launched"), ("In Training", "70% participation"),
    ("First Reflection", "Report reviewed"), ("Change in Motion", "Action plan approved"),
    ("Second Reflection", "Checkpoint 1"), ("Tempo Runner", "Checkpoint 2"),
    ("Race Ready", "Checkpoint 3"), ("Finisher", "Finish line"), ("PACE Setter", "Full journey"),
]
DESTINATIONS = ["Gold Medal", "Silver Medal", "Bronze Medal", "Personal Best",
                "Course Record", "Team Trophy", "Season Champion", "Marathon Finisher"]
COLOUR_HEX = {"Navy": "#1F3864", "Sky Blue": "#4472C4", "Emerald": "#548235", "Coral": "#ED7D31",
              "Charcoal": "#404040", "Silver": "#A6A6A6", "Gold": "#BF9000", "Teal": "#2E75B6",
              "Crimson": "#C00000", "Amber": "#FFB300"}


@dataclass
class DepartmentJourney:
    dept_id: str; dept_name: str; dept_code: str
    expected: int; actual: int; completion_pct: float
    runner_type: str; runner_colour: str; runner_status: str
    stage: str
    report_reviewed: bool; action_plan_submitted: bool
    checkpoint_1: bool; checkpoint_2: bool; checkpoint_3: bool
    destination_assigned: bool
    last_updated: str; admin_notes: str
    destination: str = ""

    @property
    def stage_index(self) -> int: return STAGE_TO_INDEX.get(self.stage, 0)
    @property
    def is_activated(self) -> bool: return self.completion_pct >= ACTIVATION_THRESHOLD
    @property
    def stage_display(self) -> str: return STAGE_DISPLAY.get(self.stage, self.stage)


def _parse_bool(v):
    if isinstance(v, bool): return v
    return str(v).strip().lower() in ("yes", "y", "true", "1")


def _yn(b): return "Yes" if b else "No"


def load_lookup(path):
    df = pd.read_excel(str(path), sheet_name="Department Lookup")
    df = df.dropna(subset=["Department ID"])
    depts = []
    for _, row in df.iterrows():
        exp = int(row["Expected Respondents"]) if not pd.isna(row["Expected Respondents"]) else 0
        act = int(row["Actual Responses"]) if not pd.isna(row["Actual Responses"]) else 0
        pct = (act / exp) if exp > 0 else 0.0
        depts.append(DepartmentJourney(
            dept_id=str(row["Department ID"]).strip(),
            dept_name=str(row["Department Name"]).strip(),
            dept_code=str(row.get("Department Code", "")).strip(),
            expected=exp, actual=act, completion_pct=round(pct, 4),
            runner_type=str(row.get("Runner Type", "Sprinter")).strip() or "Sprinter",
            runner_colour=str(row.get("Runner Colour", "Navy")).strip() or "Navy",
            runner_status=str(row.get("Runner Status", "")).strip(),
            stage=str(row.get("Current Stage", "Not Started")).strip(),
            report_reviewed=_parse_bool(row.get("Report Reviewed", "No")),
            action_plan_submitted=_parse_bool(row.get("Action Plan Submitted", "No")),
            checkpoint_1=_parse_bool(row.get("Checkpoint 1 Completed", "No")),
            checkpoint_2=_parse_bool(row.get("Checkpoint 2 Completed", "No")),
            checkpoint_3=_parse_bool(row.get("Checkpoint 3 Completed", "No")),
            destination_assigned=_parse_bool(row.get("Finish Line Reached", "No")),
            last_updated=str(row.get("Last Updated Date", "")).strip(),
            admin_notes=str(row.get("Admin Notes", "")).strip(),
        ))
    return depts


def save_lookup(path, depts):
    wb = load_workbook(str(path))
    ws = wb["Department Lookup"]
    existing = {}
    for r in range(2, ws.max_row + 1):
        did = ws.cell(row=r, column=1).value
        if did: existing[str(did).strip()] = r
    for d in depts:
        r = existing.get(d.dept_id, ws.max_row + 1)
        ws.cell(row=r, column=1, value=d.dept_id); ws.cell(row=r, column=2, value=d.dept_name)
        ws.cell(row=r, column=3, value=d.dept_code); ws.cell(row=r, column=4, value=d.expected)
        ws.cell(row=r, column=5, value=d.actual)
        ws.cell(row=r, column=7, value=d.runner_type); ws.cell(row=r, column=8, value=d.runner_colour)
        ws.cell(row=r, column=10, value=d.stage)
        ws.cell(row=r, column=11, value=_yn(d.report_reviewed)); ws.cell(row=r, column=12, value=_yn(d.action_plan_submitted))
        ws.cell(row=r, column=13, value=_yn(d.checkpoint_1)); ws.cell(row=r, column=14, value=_yn(d.checkpoint_2))
        ws.cell(row=r, column=15, value=_yn(d.checkpoint_3)); ws.cell(row=r, column=16, value=_yn(d.destination_assigned))
        ws.cell(row=r, column=17, value=datetime.now().strftime("%Y-%m-%d")); ws.cell(row=r, column=18, value=d.admin_notes)
    wb.save(str(path))


def compute_stage(d):
    if d.destination_assigned and d.checkpoint_1 and d.checkpoint_2 and d.checkpoint_3:
        return "Finish Line"
    if d.checkpoint_3: return "Race up your PACE"
    if d.checkpoint_2: return "Training - Tempo"
    if d.checkpoint_1: return "Reflect Again"
    if d.action_plan_submitted: return "Implement Change"
    if d.report_reviewed: return "Reflect"
    if d.completion_pct >= ACTIVATION_THRESHOLD: return "Training"
    if d.actual > 0 or d.expected > 0: return "Warm-up Exercise"
    return "Not Started"


def advance_stage(d, action):
    flags = {"mark_report_reviewed": ("report_reviewed", True), "approve_action_plan": ("action_plan_submitted", True),
             "mark_checkpoint_1": ("checkpoint_1", True), "mark_checkpoint_2": ("checkpoint_2", True),
             "mark_checkpoint_3": ("checkpoint_3", True), "assign_destination": ("destination_assigned", True)}
    if action in flags:
        setattr(d, flags[action][0], flags[action][1])
    d.stage = compute_stage(d)
    return d


def badges_earned(d):
    e = []
    if d.expected > 0 or d.actual > 0: e.append("Warmed Up")
    if d.is_activated: e.append("In Training")
    if d.report_reviewed: e.append("First Reflection")
    if d.action_plan_submitted: e.append("Change in Motion")
    if d.checkpoint_1: e.append("Second Reflection")
    if d.checkpoint_2: e.append("Tempo Runner")
    if d.checkpoint_3: e.append("Race Ready")
    if d.destination_assigned: e.append("Finisher")
    if d.checkpoint_1 and d.checkpoint_2 and d.checkpoint_3 and d.destination_assigned: e.append("PACE Setter")
    return e


def checkpoint_challenges(weakest_pillar=None):
    hint = {"Purpose": "This is your priority element — anchor your update here.",
            "Alliance": "Stakeholder trust and leadership visibility may be your biggest headwind.",
            "Collaboration": "Ways of working and decision-making are your primary drag.",
            "Excellence": "Learning and continuous improvement need to be more systematic."}.get(weakest_pillar, "")
    return {
        "Reflect — Purpose Alignment": {
            "focus": "Clarity, direction, and alignment.",
            "challenge": "Pacing check: some runners may not see how their leg connects to the race strategy.",
            "priority_hint": hint if weakest_pillar == "Purpose" else "",
            "prompt": "Update on how you've strengthened purpose/direction."},
        "Training Tempo — Alliance & Collaboration": {
            "focus": "Leadership, relationships, ways of working.",
            "challenge": "Tempo check: hand-offs between team-mates need cleaner exchanges to hold the pace.",
            "priority_hint": hint if weakest_pillar in ("Alliance", "Collaboration") else "",
            "prompt": "Update on alliance/collaboration."},
        "Race up your PACE — Excellence Momentum": {
            "focus": "Learning, feedback, continuous improvement.",
            "challenge": "Final surge: sustained learning and feedback are needed to hold your PACE to the line.",
            "priority_hint": hint if weakest_pillar == "Excellence" else "",
            "prompt": "Update on improvements."},
    }


def assign_random_destination(seed=None):
    rng = random.Random(seed) if seed else random
    return rng.choice(DESTINATIONS)


def log_admin_action(path, admin_user, action, dept, details):
    wb = load_workbook(str(path))
    if "Audit Log" not in wb.sheetnames: return
    ws = wb["Audit Log"]; r = ws.max_row + 1
    ws.cell(row=r, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    ws.cell(row=r, column=2, value=admin_user); ws.cell(row=r, column=3, value=action)
    ws.cell(row=r, column=4, value=dept); ws.cell(row=r, column=5, value=details)
    wb.save(str(path))


def save_action_plan(base_dir, dept_id, plan):
    d = Path(base_dir); d.mkdir(parents=True, exist_ok=True)
    out = d / f"action_plan_{dept_id}.json"
    plan["submitted_at"] = datetime.now().isoformat(timespec="seconds")
    out.write_text(json.dumps(plan, indent=2, ensure_ascii=False))
    return str(out)


def load_action_plan(base_dir, dept_id):
    f = Path(base_dir) / f"action_plan_{dept_id}.json"
    return json.loads(f.read_text()) if f.exists() else None


def save_checkpoint_update(base_dir, dept_id, checkpoint, text):
    d = Path(base_dir); d.mkdir(parents=True, exist_ok=True)
    out = d / f"checkpoints_{dept_id}.json"
    data = json.loads(out.read_text()) if out.exists() else {}
    data[f"checkpoint_{checkpoint}"] = {"updated_at": datetime.now().isoformat(timespec="seconds"), "update_text": text}
    out.write_text(json.dumps(data, indent=2, ensure_ascii=False))


def load_checkpoints(base_dir, dept_id):
    f = Path(base_dir) / f"checkpoints_{dept_id}.json"
    return json.loads(f.read_text()) if f.exists() else {}
