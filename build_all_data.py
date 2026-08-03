"""Build all Excel data files for v3.1 (PACE) — 34 departments from the attached list."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta, date
import random

random.seed(21)
NAVY = "1F3864"; GOLD_LIGHT = "FFF2CC"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _header(cell, text):
    cell.value = text
    cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border


def _data(cell, wrap=True, bold=False):
    cell.font = Font(name="Arial", size=11, bold=bold)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = border


QUESTIONS = [
    ("Q01", "Purpose", "Our team has a clear and compelling vision that guides our work.", "Vision"),
    ("Q02", "Purpose", "I understand how my role contributes to the organization's overall mission.", "Mission"),
    ("Q03", "Purpose", "Our strategic priorities are clearly communicated and consistently reinforced.", "Priorities"),
    ("Q04", "Purpose", "Our department maintains a strong focus on serving the needs of our customers.", "Customer focus"),
    ("Q05", "Purpose", "Team members share common goals and work together toward the same outcomes.", "Shared goals"),
    ("Q06", "Purpose", "My daily work is meaningfully aligned with the organization's strategic objectives.", "Alignment"),
    ("Q07", "Purpose", "Leadership regularly connects our team's work to broader business outcomes.", "Line-of-sight"),
    ("Q08", "Purpose", "Our team's purpose is well understood by our internal and external stakeholders.", "Clarity"),
    ("Q09", "Purpose", "When priorities change, leadership clearly explains the reasons and the new direction.", "Change"),
    ("Q10", "Purpose", "I feel a strong sense of purpose and meaning in the work I do here.", "Meaning"),
    ("Q11", "Alliance", "Our leaders demonstrate the behaviors expected of a high-performing culture.", "Leadership"),
    ("Q12", "Alliance", "Our team manages stakeholder relationships effectively and professionally.", "Stakeholders"),
    ("Q13", "Alliance", "Relationships within our team are characterized by mutual respect and trust.", "Trust"),
    ("Q14", "Alliance", "Cross-functional collaboration between departments is effective and productive.", "Cross-functional"),
    ("Q15", "Alliance", "Internal communication within our team is timely, clear, and consistent.", "Internal comms"),
    ("Q16", "Alliance", "External communication with customers, partners, and stakeholders is well-managed.", "External comms"),
    ("Q17", "Alliance", "I feel psychologically safe to speak up, raise concerns, and challenge ideas.", "Psych safety"),
    ("Q18", "Alliance", "Conflicts within our team are addressed constructively and resolved in a timely manner.", "Conflict"),
    ("Q19", "Alliance", "Leaders actively listen and act on feedback from team members.", "Listening"),
    ("Q20", "Alliance", "Our team consistently builds and sustains strong working partnerships with others.", "Partnerships"),
    ("Q21", "Collaboration", "Our internal processes and workflows enable us to deliver work efficiently.", "Internal"),
    ("Q22", "Collaboration", "Processes involving external parties work smoothly.", "External"),
    ("Q23", "Collaboration", "The systems we use support our work effectively and reliably.", "Systems"),
    ("Q24", "Collaboration", "We have the right tools to perform our roles at a high standard.", "Tools"),
    ("Q25", "Collaboration", "Our equipment and infrastructure meet the needs of our work.", "Equipment"),
    ("Q26", "Collaboration", "Our ways of working are well-defined, understood, and consistently applied.", "Ways of working"),
    ("Q27", "Collaboration", "Decisions in our team are made in a timely, transparent, and effective manner.", "Decisions"),
    ("Q28", "Collaboration", "Governance mechanisms clearly define roles, responsibilities, and accountabilities.", "Governance"),
    ("Q29", "Collaboration", "When processes are not working, we have effective ways to identify and improve them.", "Improvement"),
    ("Q30", "Collaboration", "Policies and controls in our team are practical and add value.", "Policy"),
    ("Q31", "Excellence", "Our team invests meaningfully in learning and development opportunities.", "L&D"),
    ("Q32", "Excellence", "The training available equips me with the skills I need to perform at a high level.", "Training"),
    ("Q33", "Excellence", "A growth mindset is actively encouraged and modeled in our team.", "Growth"),
    ("Q34", "Excellence", "We continuously look for ways to improve how we work.", "Continuous"),
    ("Q35", "Excellence", "Our team actively builds the new capabilities needed to meet future business needs.", "Capability"),
    ("Q36", "Excellence", "Innovation and new ideas are encouraged, tested, and adopted where appropriate.", "Innovation"),
    ("Q37", "Excellence", "Feedback is given and received openly, constructively, and regularly.", "Feedback"),
    ("Q38", "Excellence", "Underperformance is addressed fairly and consistently to raise overall standards.", "Perf mgmt"),
    ("Q39", "Excellence", "We celebrate achievements and recognize high performance in meaningful ways.", "Recognition"),
    ("Q40", "Excellence", "Our team is committed to continuously raising performance standards.", "Standards"),
]

# 34 departments (from the attached Departments sheet): (id, name, parent)
DEPARTMENTS = [
    ("D01", "APD - Airport Delivery", "Operations"),
    ("D02", "ASD - Airport Service Delivery", "Operations"),
    ("D03", "BIM - Brand, Insights & Marketing Communications", "Corporate"),
    ("D04", "CAF - Corporate Affairs", "Corporate"),
    ("D05", "CAY - Cathay Academy", "Corporate"),
    ("D06", "CCD - Customer Care", "Corporate"),
    ("D07", "CED - Customer Experience Design", "Operations"),
    ("D08", "CGO - Cargo", "Commercial"),
    ("D09", "CRR - Customer Relationship & Retail", "Commercial"),
    ("D10", "DEX - Digital Experience", "Operations"),
    ("D11", "DGT - Digital", "Operations"),
    ("D12", "ENG - Engineering", "Operations"),
    ("D13", "FIN - Finance", "Corporate"),
    ("D14", "FOP - Flight Operations", "Operations"),
    ("D15", "GBS - Global Business Service", "Corporate"),
    ("D16", "GIA - Group Internal Audit", "Corporate"),
    ("D17", "GLC - Group Legal & Compliance", "Corporate"),
    ("D18", "GMD - Group Medical", "Corporate"),
    ("D19", "GOR - Group Opportunities and Risks", "Corporate"),
    ("D20", "GSR - Group Safety and Operational Risk Management", "Corporate"),
    ("D21", "GSD - Group Sustainability", "Corporate"),
    ("D22", "IMT - Information Technology", "Operations"),
    ("D23", "IOC - Integrated Operations", "Operations"),
    ("D24", "ISD - Inflight Service Delivery", "Operations"),
    ("D25", "LDC - Logistics and Distribution Centre", "Commercial"),
    ("D26", "PLN - Planning", "Commercial"),
    ("D27", "PPL - People - Employee Experience", "Corporate"),
    ("D28", "PSD - Property and Service", "Operations"),
    ("D29", "PVO - Procurement & Aircraft Trading", "Operations"),
    ("D30", "REV - Revenue Management", "Commercial"),
    ("D31", "SND - Sales & Distribution", "Commercial"),
    ("D32", "OPN - Operations", "Operations"),
    ("D33", "SUB - Subsidiaries", "Operations"),
    ("D34", "EXO - Executive Team", "Corporate"),
]

def _code(name): return name.split(" - ")[0].strip()

CONFIG_SETTINGS = [
    ("tool_version", "3.1", "Tool version.", "Read-only"),
    ("schema_version", "HPC-CONFIG-v2", "Schema.", "Read-only"),
    ("framework", "PACE", "Purpose · Alliance · Collaboration · Excellence.", "Read-only"),
    ("last_updated_date", date.today().isoformat(), "Auto-stamped.", "Read-only"),
    ("last_updated_by", "Cathay Academy", "Owner.", "Editable"),
    ("scale_min", 1, "Min rating.", "Editable"),
    ("scale_max", 10, "Max rating.", "Editable"),
    ("scale_label_min", "Strongly Disagree / Very Poor", "Low label.", "Editable"),
    ("scale_label_max", "Strongly Agree / Excellent", "High label.", "Editable"),
    ("band_dysfunctional_max", 3.99, "Dysfunc upper.", "Editable"),
    ("band_balanced_max", 5.99, "Balanced upper.", "Editable"),
    ("band_performing_max", 7.99, "Performing upper.", "Editable"),
    ("imbalance_moderate_max", 2.00, "Downgrade threshold.", "Editable"),
    ("imbalance_wellbalanced_max", 1.00, "Well-balanced ceiling.", "Editable"),
    ("min_responses_dept", 10, "Min responses.", "Editable"),
    ("anonymous_mode", "TRUE", "Anonymous.", "Editable"),
    ("report_title", "High Performance Diagnostic Report", "Title.", "Editable"),
]


def build_config():
    wb = Workbook()
    ws0 = wb.active; ws0.title = "Instructions"
    ws0["A1"] = "High Performance Diagnostic Tool — Master Configuration (PACE)"
    ws0["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws0["A1"].fill = PatternFill("solid", start_color=NAVY); ws0.merge_cells("A1:F1")
    ws0["A2"] = "Powered by Cathay Academy. Elements: Purpose · Alliance · Collaboration · Excellence."
    ws0["A2"].font = Font(name="Arial", size=11, italic=True, color="595959"); ws0.merge_cells("A2:F2")
    ws = wb.create_sheet("Question Bank")
    for j, h in enumerate(["Question ID", "Pillar", "Question Text", "Active / Inactive", "Display Order", "Sub-driver", "Optional Notes"], start=1):
        _header(ws.cell(row=1, column=j), h)
    for i, (qid, pillar, text, note) in enumerate(QUESTIONS, start=2):
        ws.cell(row=i, column=1, value=qid); ws.cell(row=i, column=2, value=pillar)
        ws.cell(row=i, column=3, value=text); ws.cell(row=i, column=4, value="Active")
        ws.cell(row=i, column=5, value=i - 1); ws.cell(row=i, column=6, value=note)
        for j in range(1, 8): _data(ws.cell(row=i, column=j))
    dv = DataValidation(type="list", formula1='"Active,Inactive"', allow_blank=False); dv.add(f"D2:D{len(QUESTIONS)+1}"); ws.add_data_validation(dv)
    dv2 = DataValidation(type="list", formula1='"Purpose,Alliance,Collaboration,Excellence"', allow_blank=False); dv2.add(f"B2:B{len(QUESTIONS)+1}"); ws.add_data_validation(dv2)
    for col, w in {"A": 14, "B": 16, "C": 66, "D": 18, "E": 14, "F": 24, "G": 30}.items(): ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wsd = wb.create_sheet("Departments")
    for j, h in enumerate(["Dept ID", "Department Name", "Parent Group", "Active / Inactive", "Display Order", "HR / Business Owner", "Notes"], start=1):
        _header(wsd.cell(row=1, column=j), h)
    for i, (did, name, parent) in enumerate(DEPARTMENTS, start=2):
        vals = (did, name, parent, "Active", i - 1, "", "")
        for j, v in enumerate(vals, start=1):
            wsd.cell(row=i, column=j, value=v); _data(wsd.cell(row=i, column=j))
    dvd = DataValidation(type="list", formula1='"Active,Inactive"', allow_blank=False); dvd.add(f"D2:D{len(DEPARTMENTS)+1}"); wsd.add_data_validation(dvd)
    for col, w in {"A": 10, "B": 46, "C": 16, "D": 16, "E": 12, "F": 24, "G": 30}.items(): wsd.column_dimensions[col].width = w
    wsd.freeze_panes = "A2"
    wsc = wb.create_sheet("Config")
    for j, h in enumerate(["Setting Key", "Value", "Description", "Governance"], start=1): _header(wsc.cell(row=1, column=j), h)
    for i, (k, v, desc, gov) in enumerate(CONFIG_SETTINGS, start=2):
        wsc.cell(row=i, column=1, value=k); wsc.cell(row=i, column=2, value=v); wsc.cell(row=i, column=3, value=desc); wsc.cell(row=i, column=4, value=gov)
        for j in range(1, 5): _data(wsc.cell(row=i, column=j))
        if gov == "Read-only":
            for j in range(1, 5): wsc.cell(row=i, column=j).fill = PatternFill("solid", start_color=GOLD_LIGHT)
    for col, w in {"A": 30, "B": 40, "C": 60, "D": 16}.items(): wsc.column_dimensions[col].width = w
    wscl = wb.create_sheet("Change Log")
    for j, h in enumerate(["Date", "Changed By", "Change Type", "Item Affected", "Summary of Change", "Reason / Approval"], start=1): _header(wscl.cell(row=1, column=j), h)
    wscl.cell(row=2, column=1, value=date.today().isoformat()); wscl.cell(row=2, column=2, value="Cathay Academy")
    wscl.cell(row=2, column=3, value="Initial"); wscl.cell(row=2, column=4, value="All"); wscl.cell(row=2, column=5, value="Baseline v3.1 — 34 departments")
    for j in range(1, 7): _data(wscl.cell(row=2, column=j))
    wss = wb.create_sheet("_SCHEMA"); wss["A2"] = "schema_name"; wss["B2"] = "HPC-CONFIG-v2"; wss.sheet_state = "hidden"
    wb.save("/home/claude/hpc_tool/data/HPC_Question_Bank_Template.xlsx")
    print("Config saved (34 departments).")


PILLAR_QS = {}
for qid, pillar, text, _ in QUESTIONS:
    PILLAR_QS.setdefault(pillar, []).append((qid, text))

# Elements to render polarised (bimodal) per dept for realism
POLARISED = {"IMT - Information Technology": {"Collaboration", "Alliance"},
             "IOC - Integrated Operations": {"Collaboration"},
             "ENG - Engineering": {"Alliance"}}


def _profile(name, parent):
    """Deterministic-ish PACE profile per department."""
    base = {"Corporate": 7.0, "Operations": 6.6, "Commercial": 6.9}.get(parent, 6.8)
    r = random.Random(hash(name) & 0xffff)
    return {
        "Purpose": max(3.5, min(9.0, base + r.uniform(-0.8, 1.3))),
        "Alliance": max(3.5, min(9.0, base + r.uniform(-1.2, 1.0))),
        "Collaboration": max(3.2, min(8.8, base + r.uniform(-1.5, 0.9))),
        "Excellence": max(3.8, min(9.2, base + r.uniform(-0.6, 1.3))),
    }


def build_responses(per_dept=14):
    wb = Workbook()
    ws0 = wb.active; ws0.title = "Instructions"; ws0["A1"] = "PACE Response Data"
    ws = wb.create_sheet("Responses")
    for j, h in enumerate(["Submission ID", "Submission Timestamp", "Department", "Respondent ID", "Question ID", "Question Text", "Pillar", "Score"], start=1):
        _header(ws.cell(row=1, column=j), h)

    def clamp(x): return max(1, min(10, int(round(x))))
    all_q = [(p, q, tx) for p, qs in PILLAR_QS.items() for (q, tx) in qs]
    base_dt = datetime(2026, 7, 1, 9, 0, 0)
    row = 2; cnt = 1
    for did, name, parent in DEPARTMENTS:
        prof = _profile(name, parent)
        for r_i in range(per_dept):
            sid = f"HPC-2026-{cnt:05d}"; cnt += 1
            ts = base_dt + timedelta(hours=cnt, minutes=random.randint(0, 59))
            rid = f"EMP-{random.randint(10000, 99999)}"; po = random.gauss(0, 0.6)
            for pillar, qid, qtext in all_q:
                if name in POLARISED and pillar in POLARISED[name]:
                    s = clamp(random.gauss(prof[pillar] - 2.2, 1.0)) if r_i % 2 == 0 else clamp(random.gauss(prof[pillar] + 2.2, 1.0))
                else:
                    s = clamp(random.gauss(prof[pillar] + po, 1.4))
                ws.cell(row=row, column=1, value=sid); ws.cell(row=row, column=2, value=ts.strftime("%Y-%m-%d %H:%M:%S"))
                ws.cell(row=row, column=3, value=name); ws.cell(row=row, column=4, value=rid)
                ws.cell(row=row, column=5, value=qid); ws.cell(row=row, column=6, value=qtext)
                ws.cell(row=row, column=7, value=pillar); ws.cell(row=row, column=8, value=s)
                row += 1
    for col, w in {"A": 18, "B": 20, "C": 40, "D": 14, "E": 12, "F": 60, "G": 16, "H": 8}.items(): ws.column_dimensions[col].width = w
    wb.save("/home/claude/hpc_tool/data/HPC_Response_Data_Template.xlsx")
    print(f"Responses saved ({row - 1} rows across {len(DEPARTMENTS)} depts).")


def build_dept_lookup():
    wb = Workbook()
    ws0 = wb.active; ws0.title = "Instructions"
    ws0["A1"] = "PACE — Department Lookup"
    ws0["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws0["A1"].fill = PatternFill("solid", start_color=NAVY); ws0.merge_cells("A1:F1")
    ws0["A2"] = "Drives the PACE journey. Runner activates (In Training) at 70% completion."; ws0.merge_cells("A2:F2")
    ws = wb.create_sheet("Department Lookup")
    headers = ["Department ID", "Department Name", "Department Code", "Expected Respondents", "Actual Responses",
               "Completion %", "Runner Type", "Runner Colour", "Runner Status", "Current Stage",
               "Report Reviewed", "Action Plan Submitted", "Checkpoint 1 Completed", "Checkpoint 2 Completed",
               "Checkpoint 3 Completed", "Finish Line Reached", "Last Updated Date", "Admin Notes"]
    for j, h in enumerate(headers, start=1): _header(ws.cell(row=1, column=j), h)

    colours = list({"Navy", "Sky Blue", "Emerald", "Coral", "Charcoal", "Silver", "Gold", "Teal", "Crimson", "Amber"})
    colours = ["Navy", "Sky Blue", "Emerald", "Coral", "Charcoal", "Silver", "Gold", "Teal", "Crimson", "Amber"]
    runners = ["Sprinter", "Middle-Dist.", "Marathoner", "Relay Runner"]
    # spread of completion/stage patterns
    pattern = [
        (0.55, "Warm-up Exercise", (False, False, False, False, False, False)),
        (0.72, "Training", (False, False, False, False, False, False)),
        (0.85, "Reflect", (True, False, False, False, False, False)),
        (0.90, "Implement Change", (True, True, False, False, False, False)),
        (0.92, "Reflect Again", (True, True, True, False, False, False)),
        (0.95, "Training - Tempo", (True, True, True, True, False, False)),
    ]
    r = random.Random(7)
    for i, (did, name, parent) in enumerate(DEPARTMENTS, start=2):
        code = _code(name)
        expected = r.choice([45, 60, 80, 95, 110, 140, 180, 220])
        if name == "CAY - Cathay Academy":
            pct, stage, flags = 0.93, "Finish Line", (True, True, True, True, True, True)
        else:
            pct, stage, flags = pattern[(i - 2) % len(pattern)]
            pct += r.uniform(-0.05, 0.05)
        actual = max(0, int(round(expected * pct)))
        rr, ap, c1, c2, c3, da = flags
        ws.cell(row=i, column=1, value=did); ws.cell(row=i, column=2, value=name); ws.cell(row=i, column=3, value=code)
        ws.cell(row=i, column=4, value=expected); ws.cell(row=i, column=5, value=actual)
        ws.cell(row=i, column=6, value=f"=IF(D{i}=0,0,E{i}/D{i})"); ws.cell(row=i, column=6).number_format = "0.0%"
        ws.cell(row=i, column=7, value=runners[(i - 2) % len(runners)])
        ws.cell(row=i, column=8, value=colours[(i - 2) % len(colours)])
        ws.cell(row=i, column=9, value=f'=IF(D{i}=0,"Not set",IF(F{i}>=0.7,"In Training",IF(F{i}>=0.3,"Warming up","Not started")))')
        ws.cell(row=i, column=10, value=stage)
        for cidx, flag in zip(range(11, 17), (rr, ap, c1, c2, c3, da)):
            ws.cell(row=i, column=cidx, value="Yes" if flag else "No")
        ws.cell(row=i, column=17, value=date.today().isoformat())
        ws.cell(row=i, column=18, value="")
        for j in range(1, 19): _data(ws.cell(row=i, column=j))

    widths = {"A": 12, "B": 46, "C": 10, "D": 14, "E": 12, "F": 12, "G": 14, "H": 14, "I": 14,
              "J": 20, "K": 14, "L": 16, "M": 16, "N": 16, "O": 16, "P": 16, "Q": 14, "R": 30}
    for col, w in widths.items(): ws.column_dimensions[col].width = w
    ws.freeze_panes = "C2"
    stage_list = "Not Started,Warm-up Exercise,Training,Reflect,Implement Change,Reflect Again,Training - Tempo,Race up your PACE,Finish Line"
    dv = DataValidation(type="list", formula1=f'"{stage_list}"', allow_blank=True); dv.add(f"J2:J{len(DEPARTMENTS)+1}"); ws.add_data_validation(dv)
    dvr = DataValidation(type="list", formula1='"Sprinter,Middle-Dist.,Marathoner,Relay Runner"', allow_blank=True); dvr.add(f"G2:G{len(DEPARTMENTS)+1}"); ws.add_data_validation(dvr)
    for col in ["K", "L", "M", "N", "O", "P"]:
        d = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True); d.add(f"{col}2:{col}{len(DEPARTMENTS)+1}"); ws.add_data_validation(d)
    wsl = wb.create_sheet("Audit Log")
    for j, h in enumerate(["Timestamp", "Admin User", "Action", "Department", "Details"], start=1): _header(wsl.cell(row=1, column=j), h)
    wsl.cell(row=2, column=1, value=date.today().isoformat()); wsl.cell(row=2, column=2, value="Cathay Academy")
    wsl.cell(row=2, column=3, value="Initial"); wsl.cell(row=2, column=4, value="All"); wsl.cell(row=2, column=5, value="34 depts loaded")
    for j in range(1, 6): _data(wsl.cell(row=2, column=j))
    for col, w in {"A": 22, "B": 24, "C": 26, "D": 40, "E": 40}.items(): wsl.column_dimensions[col].width = w
    wss = wb.create_sheet("_SCHEMA"); wss["A2"] = "schema_name"; wss["B2"] = "PACE-DEPTLOOKUP-v1"; wss.sheet_state = "hidden"
    wb.save("/home/claude/hpc_tool/data/PACE_Department_Lookup.xlsx")
    print("Lookup saved (34 departments).")


if __name__ == "__main__":
    build_config(); build_responses(); build_dept_lookup()
