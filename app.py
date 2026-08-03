"""High Performance Diagnostic Tool v3 — PACE framework. Powered by Cathay Academy."""
from __future__ import annotations
import shutil
import uuid
from datetime import datetime, date
from pathlib import Path

import pandas as pd
import streamlit as st

from hpc.config_loader import (load_config, diff_configs, append_change_log,
                               HPCConfig, PILLARS, BRAND)
from hpc.engine import (analyze, load_responses, append_submission,
                        classify_score, apply_imbalance_downgrade)
from hpc import charts
from hpc.report import build_pdf
from hpc.gamification.pages import (
    page_pace, page_submit_action_plan,
    page_checkpoint_update, page_pace_admin,
)

st.set_page_config(page_title="High Performance Diagnostic Tool", page_icon="🏃",
                    layout="wide", initial_sidebar_state="expanded")

APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
CONFIG_PATH = DATA_DIR / "HPC_Question_Bank_Template.xlsx"
RESPONSE_PATH = DATA_DIR / "HPC_Response_Data_Template.xlsx"
LOOKUP_PATH = DATA_DIR / "PACE_Department_Lookup.xlsx"
GENERATED_DIR = DATA_DIR / "generated"; GENERATED_DIR.mkdir(exist_ok=True)
ACTION_PLAN_DIR = DATA_DIR / "action_plans"; ACTION_PLAN_DIR.mkdir(exist_ok=True)

NAVY = "#1F3864"; GOLD = "#BF9000"

st.markdown(f"""<style>
    .main .block-container {{ padding-top: 1.5rem; max-width: 1400px; }}
    h1, h2, h3 {{ color: {NAVY}; }}
    .stButton>button {{ background-color: {NAVY}; color: white; border: 0;
        padding: 0.5rem 1.2rem; border-radius: 4px; font-weight: 600; }}
    .stButton>button:hover {{ background-color: #2A4A7F; color: white; }}
    .hpc-hero {{ background: linear-gradient(135deg, {NAVY} 0%, #2A4A7F 100%);
        color: white; padding: 1.6rem 2rem; border-radius: 8px; margin-bottom: 1.5rem;
        border-bottom: 3px solid {GOLD}; }}
    .hpc-hero h1 {{ color: white; margin: 0; font-size: 1.8rem; }}
    .hpc-hero p  {{ color: #D9E2F3; margin: 0.4rem 0 0 0; }}
    .band-chip {{ display: inline-block; padding: 4px 10px; border-radius: 12px;
        font-size: 0.85rem; font-weight: 600; color: white; }}
    .band-dysf {{ background: #C00000; }} .band-bal {{ background: #ED7D31; }}
    .band-perf {{ background: #4472C4; }} .band-hpc {{ background: #548235; }}
    .rec-card {{ border:1px solid #E3E8F0; border-left:5px solid var(--pc);
        border-radius:6px; padding:10px 14px; margin-bottom:10px; background:#FCFDFF; }}
    .rec-card .el {{ font-weight:700; color:{NAVY}; }}
    .rec-card .pri {{ float:right; font-weight:700; font-size:0.78rem; }}
    .pol-card {{ border-left:4px solid {NAVY}; background:#F4F7FC; padding:9px 13px;
        margin-bottom:9px; border-radius:4px; }}
</style>""", unsafe_allow_html=True)

_PRI_COLOUR = {"Critical": "#C00000", "High": "#ED7D31", "Medium": "#548235", "Low": "#8C8C8C"}


def band_chip(cls):
    m = {"Dysfunctional Culture": "band-dysf", "Balanced Culture": "band-bal",
         "Performing Culture": "band-perf", "High Performance Culture": "band-hpc"}
    return f'<span class="band-chip {m.get(cls, "band-perf")}">{cls}</span>'


@st.cache_data(show_spinner=False)
def _load_cfg(path, mtime): return load_config(path)


@st.cache_data(show_spinner=False)
def _load_responses(path, mtime): return load_responses(path)


def get_config():
    if not CONFIG_PATH.exists():
        st.error("Config not found."); st.stop()
    return _load_cfg(str(CONFIG_PATH), CONFIG_PATH.stat().st_mtime)


def get_responses():
    if not RESPONSE_PATH.exists():
        return pd.DataFrame(columns=["Submission ID", "Submission Timestamp", "Department",
                                       "Respondent ID", "Question ID", "Question Text", "Pillar", "Score"])
    return _load_responses(str(RESPONSE_PATH), RESPONSE_PATH.stat().st_mtime)


def get_analysis_for_dept(dept_name):
    try:
        cfg = get_config(); df = get_responses()
        if len(df) == 0 or dept_name not in df["Department"].unique():
            return None
        return analyze(df, dept_name, cfg)
    except Exception:
        return None


def clear_caches(): st.cache_data.clear()


with st.sidebar:
    st.markdown(f"<h2 style='color:{NAVY};margin:0'>🏃 PACE Diagnostic</h2>", unsafe_allow_html=True)
    st.markdown(f"<p style='color:#666;margin:0 0 1rem 0'>{BRAND}</p>", unsafe_allow_html=True)
    page = st.radio("Nav",
        ["🏠 Home", "📝 Take Questionnaire", "📊 Admin Dashboard",
         "📄 Generate Executive Report", "📥 Upload Response Data",
         "⚙️ Upload Configuration", "———",
         "🏃 PACE", "🎛️ Submit Action Plan",
         "📍 Checkpoint Update", "🛠️ PACE Admin"],
        label_visibility="collapsed")
    st.markdown("---")
    try:
        cfg = get_config()
        st.caption(f"**Questions:** {len(cfg.active_questions)}")
        st.caption(f"**Elements:** {' · '.join(PILLARS)}")
        st.caption(f"**Departments:** {len(cfg.active_departments)}")
    except Exception: pass
    try:
        n = get_responses()["Submission ID"].nunique()
        st.caption(f"**Submissions:** {n}")
    except Exception: pass


def page_home():
    st.markdown(f"""<div class="hpc-hero">
    <h1>High Performance Diagnostic Tool</h1>
    <p>Diagnostic across the <b>PACE</b> framework — <b>P</b>urpose · <b>A</b>lliance ·
    <b>C</b>ollaboration · <b>E</b>xcellence. {BRAND}.</p></div>""", unsafe_allow_html=True)
    cfg = get_config(); df = get_responses()
    if len(df) == 0:
        st.info("No response data loaded yet."); return
    n_subs = df["Submission ID"].nunique(); n_depts = df["Department"].nunique()
    cp = df.groupby("Pillar")["Score"].mean().reindex(PILLARS)
    co = float(cp.mean()); imb = float(cp.max() - cp.min()); cls = classify_score(co, cfg)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Submissions", f"{n_subs:,}")
    c2.metric("Departments", n_depts)
    c3.metric("Company PACE", f"{co:.2f}", f"{imb:.2f} gap")
    c4.markdown(f"**Classification**<br>{band_chip(cls)}", unsafe_allow_html=True)
    pv = pd.DataFrame({"Element": PILLARS, "Score": [round(float(cp[p]), 2) for p in PILLARS],
                        "Status": [classify_score(float(cp[p]), cfg) for p in PILLARS]})
    st.dataframe(pv, use_container_width=True, hide_index=True)


# ---------------------------------------------------------------------------
# QUESTIONNAIRE — no default scores; all questions required
# ---------------------------------------------------------------------------
def page_questionnaire():
    st.markdown(f"""<div class="hpc-hero"><h1>Take the Questionnaire</h1>
    <p>Anonymous · 40 questions across Purpose · Alliance · Collaboration · Excellence.
    Every question must be answered before you can submit.</p></div>""", unsafe_allow_html=True)
    cfg = get_config(); active = cfg.active_questions; depts = cfg.active_departments
    if not depts:
        st.error("No active departments."); return
    if "quest_step" not in st.session_state: st.session_state.quest_step = "start"
    if "quest_answers" not in st.session_state: st.session_state.quest_answers = {}

    if st.session_state.quest_step == "start":
        dept = st.selectbox("Select department", [""] + depts)
        st.caption(f"You will answer {len(active)} questions on a {cfg.scale_min}–{cfg.scale_max} scale "
                   f"({cfg.scale_min} = {cfg.scale_label_min}, {cfg.scale_max} = {cfg.scale_label_max}). "
                   "No answers are pre-filled — please rate every question.")
        if st.button("Start", type="primary", disabled=(dept == "")):
            st.session_state.quest_dept = dept
            st.session_state.quest_step = "answer"
            st.session_state.quest_answers = {}
            st.rerun()
        return

    if st.session_state.quest_step == "answer":
        st.markdown(f"**Department:** {st.session_state.quest_dept}")
        total = len(active); opts = list(range(cfg.scale_min, cfg.scale_max + 1))
        tabs = st.tabs(PILLARS)
        for tab, pillar in zip(tabs, PILLARS):
            with tab:
                st.caption(f"1 = {cfg.scale_label_min}  ·  10 = {cfg.scale_label_max}")
                for _, row in active[active["Pillar"] == pillar].iterrows():
                    qid = row["Question ID"]
                    prev = st.session_state.quest_answers.get(qid)
                    idx = opts.index(prev) if prev in opts else None
                    val = st.radio(f"**{qid}.** {row['Question Text']}", opts,
                                   index=idx, horizontal=True, key=f"q_{qid}")
                    if val is not None:
                        st.session_state.quest_answers[qid] = val
                    st.markdown("---")
        answered = sum(1 for _, r in active.iterrows()
                       if st.session_state.quest_answers.get(r["Question ID"]) is not None)
        st.progress(answered / total, text=f"{answered} / {total} answered")
        if total - answered > 0:
            st.warning(f"⚠️ {total - answered} question(s) still need an answer before you can submit.")
        c1, _, c3 = st.columns([1, 2, 1])
        with c1:
            if st.button("← Back"): st.session_state.quest_step = "start"; st.rerun()
        with c3:
            if st.button("Submit ✓", type="primary", disabled=(answered != total)):
                sub_id = f"HPC-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
                answers = {r["Question ID"]: (r["Question Text"], r["Pillar"],
                                                int(st.session_state.quest_answers[r["Question ID"]]))
                            for _, r in active.iterrows()}
                _ensure_response_file()
                append_submission(str(RESPONSE_PATH), sub_id, st.session_state.quest_dept,
                                    "" if cfg.anonymous_mode else f"USER-{uuid.uuid4().hex[:6].upper()}", answers)
                clear_caches()
                st.session_state.quest_step = "done"; st.session_state.quest_sub_id = sub_id; st.rerun()

    if st.session_state.quest_step == "done":
        st.balloons(); st.success("✅ Thank you — all questions completed and submitted.")
        st.markdown(f"**Submission ID:** {st.session_state.quest_sub_id}")
        if st.button("Take another"):
            st.session_state.quest_step = "start"; st.session_state.quest_answers = {}; st.rerun()


def _ensure_response_file():
    if RESPONSE_PATH.exists(): return
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Responses"
    for j, h in enumerate(["Submission ID", "Submission Timestamp", "Department", "Respondent ID",
                            "Question ID", "Question Text", "Pillar", "Score"], start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F3864")
        c.alignment = Alignment(horizontal="center")
    wb.save(RESPONSE_PATH)


# ---------------------------------------------------------------------------
# ADMIN DASHBOARD — single combined view, robust, clear individual charts
# ---------------------------------------------------------------------------
def page_dashboard():
    st.markdown(f"""<div class="hpc-hero"><h1>Admin Dashboard</h1>
    <p>PACE insights in one view — clear charts, polarisation &amp; variance analysis,
    and tailored recommendations.</p></div>""", unsafe_allow_html=True)
    cfg = get_config(); df = get_responses()
    if len(df) == 0:
        st.warning("No data."); return

    depts_avail = sorted(df["Department"].unique())
    c1, c2 = st.columns([2, 1])
    with c1:
        focus_opt = st.selectbox("Focus", ["Company-wide (all)"] + depts_avail)
    with c2:
        min_r = st.number_input("Min responses (caution flag)", min_value=1, value=cfg.min_responses_dept)

    # Robust: analyse against the FULL dataset so the focus dept is always present.
    focus_key = "__ALL__" if focus_opt.startswith("Company") else focus_opt
    try:
        analysis = analyze(df, focus_key, cfg)
    except Exception as e:
        st.error(f"Could not analyse: {e}")
        return

    if analysis.focus.n_respondents < min_r and analysis.focus.department != "Company-wide":
        st.warning(f"⚠️ {analysis.focus.department} has {analysis.focus.n_respondents} responses "
                   f"(< {min_r}); interpret with caution.")

    # ---- Focus header (small, no giant metric text) ----
    st.markdown(
        f"<div style='display:flex;gap:24px;align-items:center;flex-wrap:wrap;margin:6px 0 4px'>"
        f"<span style='font-size:1.15rem;font-weight:700;color:{NAVY}'>Focus: {analysis.focus.department}</span>"
        f"<span>Respondents: <b>{analysis.focus.n_respondents:,}</b></span>"
        f"<span>PACE score: <b>{analysis.focus.overall:.2f}</b></span>"
        f"<span>vs company: <b>{analysis.focus.overall - analysis.company_overall:+.2f}</b></span>"
        f"{band_chip(analysis.focus.classification)}</div>", unsafe_allow_html=True)

    # ---- Clear individual charts (no cramped one-page image) ----
    r1a, r1b = st.columns(2)
    with r1a:
        st.pyplot(charts.radar_chart(analysis.focus.pillar_means, analysis.company_pillar_means,
                                     analysis.focus.department), use_container_width=True)
    with r1b:
        st.pyplot(charts.pillar_bar(analysis.focus.pillar_means, analysis.company_pillar_means),
                  use_container_width=True)
    r2a, r2b = st.columns(2)
    with r2a:
        st.pyplot(charts.polarisation_bar(analysis.polarisation), use_container_width=True)
    with r2b:
        st.pyplot(charts.ranking_bar(analysis.all_departments, analysis.company_overall,
                                     focus=analysis.focus.department if analysis.focus.department != "Company-wide" else None),
                  use_container_width=True)

    st.markdown("---")

    # ---- Polarisation & Variance (varied statements) ----
    st.markdown("### 🔬 Polarisation & Variance Analysis")
    pcols = st.columns(2)
    for i, pol in enumerate(analysis.polarisation):
        badge = "🔴 Polarised" if pol.polarised else pol.variance_label
        with pcols[i % 2]:
            st.markdown(
                f"<div class='pol-card'><b>{pol.pillar}</b> "
                f"<span style='color:#8C8C8C'>· mean {pol.mean} · SD {pol.std} · {badge}</span><br>"
                f"<span style='font-size:0.92rem'>{pol.statement}</span></div>", unsafe_allow_html=True)

    # ---- Tailored recommendations as cards (no horizontal scroll) ----
    st.markdown("### 🎯 Tailored Actionable Recommendations")
    tcols = st.columns(2)
    for i, t in enumerate(analysis.tailored_recommendations):
        col = _PRI_COLOUR.get(t["Priority"], "#8C8C8C")
        with tcols[i % 2]:
            st.markdown(
                f"<div class='rec-card' style='--pc:{col}'>"
                f"<span class='el'>{t['Element']}</span>"
                f"<span class='pri' style='color:{col}'>{t['Priority']}</span><br>"
                f"<span style='font-size:0.92rem'>{t['Recommendation']}</span><br>"
                f"<span style='font-size:0.8rem;color:#667'>Mean {t['Mean']} · Gap {t['Gap']} · {t['Why']}</span>"
                f"</div>", unsafe_allow_html=True)

    # ---- Supporting detail (bullets, no wide tables) ----
    with st.expander("📋 Rule-based insights & department table"):
        for ins in analysis.insights:
            st.markdown(f"- **{ins['label']}:** {ins['text']}")
        st.markdown("**Department overview** (Overall + imbalance)")
        tbl = analysis.all_departments[PILLARS + ["Overall", "Imbalance", "N respondents"]].round(2).copy()
        cls = []
        for d in tbl.index:
            base = classify_score(tbl.loc[d, "Overall"], cfg)
            final, dg = apply_imbalance_downgrade(base, tbl.loc[d, "Imbalance"], cfg)
            cls.append(final + (" ⚠" if dg else ""))
        tbl["Classification"] = cls
        st.dataframe(tbl, use_container_width=True)


def page_upload_responses():
    st.markdown(f"""<div class="hpc-hero"><h1>Upload Response Data</h1>
    <p>Import .xlsx or .csv.</p></div>""", unsafe_allow_html=True)
    up = st.file_uploader("File", type=["xlsx", "csv"])
    mode = st.radio("Mode", ["Replace", "Append"], horizontal=True)
    if up is not None:
        tmp = DATA_DIR / f"_incoming_{up.name}"; tmp.write_bytes(up.getvalue())
        try:
            incoming = load_responses(str(tmp))
        except Exception as e:
            st.error(f"Rejected: {e}"); tmp.unlink(missing_ok=True); return
        st.success(f"✅ {len(incoming):,} rows.")
        st.dataframe(incoming.head(10), use_container_width=True)
        if st.button("Apply ✓", type="primary"):
            if mode == "Replace":
                shutil.copy(tmp, RESPONSE_PATH)
            else:
                existing = get_responses() if RESPONSE_PATH.exists() else pd.DataFrame()
                pd.concat([existing, incoming], ignore_index=True).to_excel(RESPONSE_PATH, sheet_name="Responses", index=False)
            tmp.unlink(missing_ok=True); clear_caches(); st.success("✅ Applied.")


def page_upload_config():
    st.markdown(f"""<div class="hpc-hero"><h1>Upload Configuration</h1>
    <p>Edit config Excel then upload.</p></div>""", unsafe_allow_html=True)
    current = get_config()
    c1, c2, c3 = st.columns(3)
    c1.metric("Questions", len(current.active_questions))
    c2.metric("Departments", len(current.active_departments))
    c3.metric("Version", str(current.get("tool_version", "3.0")))
    with open(CONFIG_PATH, "rb") as f:
        st.download_button("⬇️ Download config", f.read(), file_name="HPC_Question_Bank_Template.xlsx")
    st.markdown("---")
    up = st.file_uploader("Upload", type=["xlsx"])
    if up is not None:
        tmp = DATA_DIR / f"_incoming_{up.name}"; tmp.write_bytes(up.getvalue())
        try:
            incoming = load_config(str(tmp))
        except Exception as e:
            st.error(f"❌ {e}"); tmp.unlink(missing_ok=True); return
        st.success("✅ Valid.")
        st.markdown("### Diff")
        diff = diff_configs(current, incoming)
        for cat, items in diff.items():
            if items:
                with st.expander(f"{cat} ({len(items)})", expanded=len(items) < 5):
                    for item in items:
                        st.markdown(f"- {item}")
        changed_by = st.text_input("Changed by")
        change_type = st.selectbox("Type", ["Update", "Question edit", "Dept edit", "Other"])
        summary = st.text_area("Summary", height=80)
        if st.button("Apply ✓", type="primary", disabled=(not changed_by or not summary)):
            backup = DATA_DIR / f"HPC_Config.backup.{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            shutil.copy(CONFIG_PATH, backup); shutil.copy(tmp, CONFIG_PATH); tmp.unlink(missing_ok=True)
            append_change_log(str(CONFIG_PATH), [{"Date": date.today().isoformat(), "Changed By": changed_by,
                "Change Type": change_type, "Item Affected": "Multiple", "Summary of Change": summary,
                "Reason / Approval": "Via tool"}])
            clear_caches(); st.success(f"✅ Applied. Backup: {backup.name}")


def page_generate_report():
    st.markdown(f"""<div class="hpc-hero"><h1>Generate Executive Report</h1>
    <p>One-page PDF with polarisation analysis and tailored recommendations.</p></div>""", unsafe_allow_html=True)
    cfg = get_config(); df = get_responses()
    if len(df) == 0:
        st.warning("No data."); return
    focus = st.selectbox("Focus", ["Company-wide (all)"] + sorted(df["Department"].unique()))
    prepared_by = st.text_input("Prepared by", value="OD — Cathay Academy")
    dt_str = st.text_input("Date", value=datetime.now().strftime("%d %B %Y"))
    if st.button("Generate PDF ✓", type="primary"):
        with st.spinner("Building..."):
            key = "__ALL__" if focus.startswith("Company") else focus
            analysis = analyze(df, key, cfg)
            fn = f"PACE_Report_{focus.replace(' ','_').replace('(','').replace(')','')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            out = GENERATED_DIR / fn
            build_pdf(analysis, cfg, str(out), prepared_by=prepared_by, analysis_date=dt_str)
        st.success("✅ Generated (one-page).")
        with open(out, "rb") as f:
            st.download_button("📄 Download PDF", f.read(), file_name=fn, mime="application/pdf")


if page == "🏠 Home": page_home()
elif page == "📝 Take Questionnaire": page_questionnaire()
elif page == "📊 Admin Dashboard": page_dashboard()
elif page == "📄 Generate Executive Report": page_generate_report()
elif page == "📥 Upload Response Data": page_upload_responses()
elif page == "⚙️ Upload Configuration": page_upload_config()
elif page == "🏃 PACE": page_pace(str(LOOKUP_PATH), str(ACTION_PLAN_DIR))
elif page == "🎛️ Submit Action Plan":
    page_submit_action_plan(str(LOOKUP_PATH), str(ACTION_PLAN_DIR), get_analysis_for_dept)
elif page == "📍 Checkpoint Update":
    page_checkpoint_update(str(LOOKUP_PATH), str(ACTION_PLAN_DIR), get_analysis_for_dept)
elif page == "🛠️ PACE Admin":
    page_pace_admin(str(LOOKUP_PATH), str(ACTION_PLAN_DIR), str(DATA_DIR))

st.markdown(f"""<div style="margin-top:3rem;padding-top:1rem;border-top:1px solid #E7E7E7;
color:#8C8C8C;font-size:0.85rem;text-align:center">
{BRAND} · v3.0 · PACE: Purpose · Alliance · Collaboration · Excellence
</div>""", unsafe_allow_html=True)
